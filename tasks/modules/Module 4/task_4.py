from pathlib import Path
import openpyxl
from pyinputplus import inputStr


def reverse_excel_table():
    """Reverse (column/row) for table in excel file."""
    spreadsheet_path = Path(inputStr("Provide absolute path to excel file: "))
    wb = openpyxl.load_workbook(spreadsheet_path)
    ws_after = wb.create_sheet("After")
    ws_after.title = "reversed_sheet"
    for row in wb[wb.sheetnames[0]].iter_rows():
        for cell in row:
            value = cell.value
            old_row = cell.row
            old_col = cell.column
            wb[ws_after.title].cell(row=old_col, column=old_row).value = value
    wb.save(spreadsheet_path)


if __name__ == '__main__':
    reverse_excel_table()
